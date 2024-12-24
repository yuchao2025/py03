import pymysql
from openpyxl import load_workbook

def insert_excel_data_to_mysql(in_filename, in_host, in_user, in_password, in_database, in_table):
    # 连接数据库
    mydb = pymysql.connect(
        host=in_host,  # 数据库主机地址
        user=in_user,  # 数据库用户名
        password=in_password,  # 数据库密码
        database=in_database  # 数据库名称
    )

    # 创建游标对象
    cursor = mydb.cursor()

    # 打开一个 Excel 文件
    workbook = load_workbook(in_filename)
    # 选择第一个工作表
    sheet = workbook.worksheets[0]
    # # 或者获取活动工作表
    # sheet = workbook.active

    # 读取整行或整列数据
    row_values = [cell.value for cell in sheet[1]]  # 第一行数据
    # column_values = [sheet[f'A{i}'].value for i in range(1, sheet.max_row + 1)]  # A列数据

    # 获取insert语句中Values(%s,%s)中%s占位符的个数
    row_param = ['%s' for cell in sheet[1]]

    # 拼接插入数据的SQL语句
    sql = "INSERT INTO {0} {1} VALUES {2};"  # SQL插入语句
    # 将语句进行格式化，并将字段名称和占位符的单引号替换成空
    sql_format = sql.format(in_table, tuple(row_values), tuple(row_param)).replace("'", "")

    # 开始导入
    # 输出导入文件
    print("Import start from " + in_filename)
    # 输出导入表
    print("Import table " + in_table)
    # 输出SQL语句模板
    print("SQL:" + sql_format)
    # 输出导入字段
    print("导入字段：" + str(row_values))

    # 导入行数
    add_row_count = 0
    # 遍历Excel表格中的每一行，并将每一行插入到数据库中（从第2行开始）
    for row in sheet.iter_rows(values_only=True, min_row=2):
        add_row_count += 1
        cursor.execute(sql_format, row)  # 执行SQL插入语句

    # 提交更改并关闭数据库连接
    mydb.commit()  # 提交更改
    cursor.close()  # 关闭游标对象
    mydb.close()  # 关闭数据库连接

    # 打印结果提示
    print("\033[92m" + "Added rows:" + str(add_row_count) + "\033[0m")
    print("\033[92m"+"Import successfully! " + "\033[0m")


if __name__ == '__main__':
    # 需要输入的参数
    #filename = r'/root/python01/output.xlsx'  # Excel文件路径
    filename = r'D:\Downloads\output.xlsx'
    host = "172.16.31.67"  # 数据库主机地址
    user = "root"  # 数据库用户名
    password = "123@qq.com"  # 数据库密码
    database = "test1"  # 数据库名称
    table = "sys_job1"  # 数据库表名

    # 调用函数，将Excel数据插入到MySQL数据库中(Excel中第一行是表中的字段名称，第二行开始为数据)
    insert_excel_data_to_mysql(filename, host, user, password, database, table)
